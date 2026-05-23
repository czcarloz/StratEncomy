"""Parse B3 Excel exports — Negociações and Movimentações formats."""
from __future__ import annotations

import io
from dataclasses import dataclass, field
from datetime import datetime
from decimal import Decimal, InvalidOperation

import openpyxl

# Movimentações: types that map to buy/sell
_MOV_OP_TYPES = {"Transferência - Liquidação", "Compra", "Venda"}

# Movimentações: types that map to dividends
_MOV_DIV_TYPES = {"Rendimento", "Juros Sobre Capital Próprio", "Dividendo"}


@dataclass
class ParsedOperation:
    ticker: str
    op_type: str        # "buy" | "sell"
    quantity: str
    unit_price: str
    date: str           # YYYY-MM-DD
    broker: str | None


@dataclass
class ParsedDividend:
    ticker: str
    amount: str
    date: str           # YYYY-MM-DD
    note: str | None


@dataclass
class SkippedRow:
    row_num: int
    reason: str
    movimentacao: str | None


@dataclass
class B3ParseResult:
    operations: list[ParsedOperation] = field(default_factory=list)
    dividends: list[ParsedDividend] = field(default_factory=list)
    skipped: list[SkippedRow] = field(default_factory=list)
    source: str = ""   # "negociacao" | "movimentacao"


# ── Shared helpers ────────────────────────────────────────────────────────────

def _parse_date(val: str) -> str:
    """DD/MM/YYYY → YYYY-MM-DD"""
    return datetime.strptime(str(val).strip(), "%d/%m/%Y").strftime("%Y-%m-%d")


def _parse_decimal(val: object) -> Decimal | None:
    if val is None:
        return None
    s = str(val).strip()
    if s in ("-", "", "None"):
        return None
    try:
        return Decimal(s)
    except InvalidOperation:
        return None


def _abbreviate_broker(name: str) -> str:
    """'XP INVESTIMENTOS CCTVM S/A' → 'XP'"""
    parts = str(name).strip().split()
    return parts[0] if parts else str(name)


def _extract_ticker_from_produto(produto: str) -> str:
    """'PETR4 - PETROLEO BRASILEIRO S/A' → 'PETR4'"""
    return str(produto).split(" - ")[0].strip()


def _normalize_ticker(ticker: str, mercado: str) -> str:
    """Remove trailing 'F' for fractional market tickers (ITSA4F → ITSA4)."""
    t = str(ticker).strip().upper()
    if "Fracionário" in mercado and t.endswith("F"):
        t = t[:-1]
    return t


def _guess_asset_class(ticker: str) -> str:
    """Best-effort heuristic — user can change after import."""
    t = ticker.upper()
    if t.startswith("TESOURO"):
        return "bond"
    if t.endswith("11"):
        return "fii"
    if t.endswith(("3", "4", "5", "6", "7", "8")):
        return "stock"
    return "stock"


# ── Negociação parser ─────────────────────────────────────────────────────────
# Columns: Data do Negócio, Tipo de Movimentação, Mercado, Prazo/Vencimento,
#          Instituição, Código de Negociação, Quantidade, Preço, Valor

def _parse_negociacao(ws) -> B3ParseResult:
    result = B3ParseResult(source="negociacao")

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not row[0]:
            continue

        data       = str(row[0]).strip() if row[0] else ""
        tipo       = str(row[1]).strip() if row[1] else ""
        mercado    = str(row[2]).strip() if row[2] else ""
        instituicao = str(row[4]).strip() if row[4] else ""
        codigo     = str(row[5]).strip() if row[5] else ""
        quantidade = row[6]
        preco      = row[7]

        if not tipo or not codigo:
            continue

        if tipo not in ("Compra", "Venda"):
            result.skipped.append(SkippedRow(i, "Tipo ignorado", tipo))
            continue

        qty   = _parse_decimal(quantidade)
        price = _parse_decimal(preco)

        if qty is None or qty <= 0:
            result.skipped.append(SkippedRow(i, "Quantidade inválida", tipo))
            continue
        if price is None or price <= 0:
            result.skipped.append(SkippedRow(i, "Preço ausente", tipo))
            continue

        result.operations.append(ParsedOperation(
            ticker=_normalize_ticker(codigo, mercado),
            op_type="buy" if tipo == "Compra" else "sell",
            quantity=str(qty),
            unit_price=str(price),
            date=_parse_date(data),
            broker=_abbreviate_broker(instituicao) if instituicao else None,
        ))

    return result


# ── Movimentação parser ───────────────────────────────────────────────────────
# Columns: Entrada/Saída, Data, Movimentação, Produto, Instituição,
#          Quantidade, Preço unitário, Valor da Operação

def _parse_movimentacao(ws) -> B3ParseResult:
    result = B3ParseResult(source="movimentacao")

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not row[0]:
            continue

        entrada_saida = str(row[0]).strip() if row[0] else ""
        data          = str(row[1]).strip() if row[1] else ""
        movimentacao  = str(row[2]).strip() if row[2] else ""
        produto       = str(row[3]).strip() if row[3] else ""
        instituicao   = str(row[4]).strip() if row[4] else ""
        quantidade    = row[5]
        preco_unit    = row[6]
        valor_op      = row[7]

        if not movimentacao or not produto:
            continue

        if movimentacao in _MOV_OP_TYPES:
            price = _parse_decimal(preco_unit)
            qty   = _parse_decimal(quantidade)
            if price is None or price <= 0:
                result.skipped.append(SkippedRow(i, "Preço unitário ausente", movimentacao))
                continue
            if qty is None or qty <= 0:
                result.skipped.append(SkippedRow(i, "Quantidade inválida", movimentacao))
                continue

            if movimentacao == "Venda" or (
                movimentacao == "Transferência - Liquidação" and entrada_saida == "Debito"
            ):
                op_type = "sell"
            else:
                op_type = "buy"

            result.operations.append(ParsedOperation(
                ticker=_extract_ticker_from_produto(produto),
                op_type=op_type,
                quantity=str(qty),
                unit_price=str(price),
                date=_parse_date(data),
                broker=_abbreviate_broker(instituicao) if instituicao else None,
            ))

        elif movimentacao in _MOV_DIV_TYPES:
            amount = _parse_decimal(valor_op)
            if amount is None or amount <= 0:
                result.skipped.append(SkippedRow(i, "Valor ausente ou zero", movimentacao))
                continue

            result.dividends.append(ParsedDividend(
                ticker=_extract_ticker_from_produto(produto),
                amount=str(amount),
                date=_parse_date(data),
                note=movimentacao if movimentacao != "Rendimento" else None,
            ))

        else:
            result.skipped.append(SkippedRow(i, "Tipo ignorado", movimentacao))

    return result


# ── Auto-detect and parse ─────────────────────────────────────────────────────

def parse_b3_xlsx(file_bytes: bytes) -> B3ParseResult:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)

    if "Negociação" in wb.sheetnames:
        return _parse_negociacao(wb["Negociação"])

    if "Movimentação" in wb.sheetnames:
        return _parse_movimentacao(wb["Movimentação"])

    raise ValueError(
        "Arquivo não reconhecido. Faça o download do Extrato de Negociações "
        "ou Extrato de Movimentações em investidor.b3.com.br."
    )
