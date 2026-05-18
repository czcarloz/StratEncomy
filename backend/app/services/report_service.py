"""Generates PDF and XLSX reports from transaction data."""
from __future__ import annotations

import io
from dataclasses import dataclass
from decimal import Decimal

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable,
)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# ── Data contract ────────────────────────────────────────────────────────────

@dataclass
class CategoryRow:
    name: str
    amount: Decimal
    percentage: float


@dataclass
class TransactionRow:
    date: str
    description: str
    category: str
    type: str
    amount: Decimal


@dataclass
class ReportData:
    month_label: str          # e.g. "May 2026"
    total_income: Decimal
    total_expense: Decimal
    total_investment: Decimal
    balance: Decimal
    categories: list[CategoryRow]
    transactions: list[TransactionRow]


# ── Palette ──────────────────────────────────────────────────────────────────

_SLATE_900 = colors.HexColor("#0F172A")
_SLATE_700 = colors.HexColor("#334155")
_SLATE_400 = colors.HexColor("#94A3B8")
_WHITE     = colors.white
_GREEN     = colors.HexColor("#059669")
_RED       = colors.HexColor("#DC2626")
_BLUE      = colors.HexColor("#3B82F6")


def _fmt(v: Decimal) -> str:
    from babel.numbers import format_currency  # type: ignore
    try:
        return format_currency(v, "BRL", locale="pt_BR")
    except Exception:
        return f"R$ {v:,.2f}"


def _fmt_safe(v: Decimal) -> str:
    """Format without babel dependency."""
    return f"R$ {float(v):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


# ── PDF ──────────────────────────────────────────────────────────────────────

def generate_pdf(data: ReportData) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=2 * cm, rightMargin=2 * cm,
        topMargin=2 * cm, bottomMargin=2 * cm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "Title", parent=styles["Normal"],
        fontSize=18, textColor=_SLATE_900, spaceAfter=4,
        fontName="Helvetica-Bold",
    )
    subtitle_style = ParagraphStyle(
        "Subtitle", parent=styles["Normal"],
        fontSize=10, textColor=_SLATE_400, spaceAfter=12,
    )
    section_style = ParagraphStyle(
        "Section", parent=styles["Normal"],
        fontSize=11, textColor=_SLATE_900, spaceBefore=16, spaceAfter=8,
        fontName="Helvetica-Bold",
    )

    story = []

    # ── Header
    story.append(Paragraph("StratEncomy", title_style))
    story.append(Paragraph(f"Transactions Report — {data.month_label}", subtitle_style))
    story.append(HRFlowable(width="100%", thickness=1, color=_SLATE_700, spaceAfter=16))

    # ── Summary cards (4-col table)
    story.append(Paragraph("Summary", section_style))
    summary_data = [
        ["Income", "Expenses", "Invested", "Balance"],
        [
            _fmt_safe(data.total_income),
            _fmt_safe(data.total_expense),
            _fmt_safe(data.total_investment),
            _fmt_safe(data.balance),
        ],
    ]
    summary_table = Table(summary_data, colWidths=["25%"] * 4)
    summary_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), _SLATE_900),
        ("TEXTCOLOR",  (0, 0), (-1, 0), _WHITE),
        ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",   (0, 0), (-1, 0), 9),
        ("ALIGN",      (0, 0), (-1, -1), "CENTER"),
        ("VALIGN",     (0, 0), (-1, -1), "MIDDLE"),
        ("FONTNAME",   (0, 1), (-1, 1), "Helvetica-Bold"),
        ("FONTSIZE",   (0, 1), (-1, 1), 11),
        ("TEXTCOLOR",  (0, 1), (0, 1), _GREEN),
        ("TEXTCOLOR",  (1, 1), (1, 1), _RED),
        ("TEXTCOLOR",  (2, 1), (2, 1), _BLUE),
        ("TEXTCOLOR",  (3, 1), (3, 1), _GREEN if data.balance >= 0 else _RED),
        ("ROWBACKGROUNDS", (0, 1), (-1, 1), [colors.HexColor("#F8FAFC")]),
        ("BOX",  (0, 0), (-1, -1), 0.5, _SLATE_700),
        ("GRID", (0, 0), (-1, -1), 0.25, _SLATE_400),
        ("TOPPADDING",    (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
    ]))
    story.append(summary_table)

    # ── Expense by category
    if data.categories:
        story.append(Paragraph("Expenses by Category", section_style))
        cat_data = [["Category", "Amount", "%"]] + [
            [c.name, _fmt_safe(c.amount), f"{c.percentage:.1f}%"]
            for c in data.categories
        ]
        cat_table = Table(cat_data, colWidths=["55%", "30%", "15%"])
        cat_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), _SLATE_900),
            ("TEXTCOLOR",  (0, 0), (-1, 0), _WHITE),
            ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE",   (0, 0), (-1, -1), 9),
            ("ALIGN",      (1, 0), (-1, -1), "RIGHT"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [_WHITE, colors.HexColor("#F1F5F9")]),
            ("BOX",  (0, 0), (-1, -1), 0.5, _SLATE_700),
            ("LINEBELOW", (0, 0), (-1, 0), 0.5, _SLATE_700),
            ("TOPPADDING",    (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))
        story.append(cat_table)

    # ── Transactions
    story.append(Paragraph("Transactions", section_style))
    if not data.transactions:
        story.append(Paragraph("No transactions for this period.", styles["Normal"]))
    else:
        tx_data = [["Date", "Description", "Category", "Type", "Amount"]] + [
            [t.date, t.description or "—", t.category, t.type.capitalize(), _fmt_safe(t.amount)]
            for t in data.transactions
        ]
        tx_table = Table(tx_data, colWidths=["12%", "33%", "20%", "12%", "23%"])
        tx_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), _SLATE_900),
            ("TEXTCOLOR",  (0, 0), (-1, 0), _WHITE),
            ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE",   (0, 0), (-1, -1), 8),
            ("ALIGN",      (4, 0), (4, -1), "RIGHT"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [_WHITE, colors.HexColor("#F1F5F9")]),
            ("BOX",  (0, 0), (-1, -1), 0.5, _SLATE_700),
            ("LINEBELOW", (0, 0), (-1, 0), 0.5, _SLATE_700),
            ("TOPPADDING",    (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        story.append(tx_table)

    doc.build(story)
    return buf.getvalue()


# ── XLSX ─────────────────────────────────────────────────────────────────────

def generate_xlsx(data: ReportData) -> bytes:
    wb = Workbook()

    _header_font   = Font(bold=True, color="F8FAFC", size=10)
    _header_fill   = PatternFill("solid", fgColor="0F172A")
    _alt_fill      = PatternFill("solid", fgColor="F1F5F9")
    _thin_border   = Border(
        left=Side(style="thin", color="334155"),
        right=Side(style="thin", color="334155"),
        top=Side(style="thin", color="334155"),
        bottom=Side(style="thin", color="334155"),
    )

    def _style_header(cell):
        cell.font = _header_font
        cell.fill = _header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _thin_border

    def _style_cell(cell, alt=False):
        if alt:
            cell.fill = _alt_fill
        cell.alignment = Alignment(vertical="center")
        cell.border = _thin_border

    # ── Sheet 1: Summary ──────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    ws.append([f"StratEncomy — Transactions Report — {data.month_label}"])
    ws["A1"].font = Font(bold=True, size=13)
    ws.append([])
    ws.append(["Income", "Expenses", "Invested", "Balance"])
    ws.append([
        float(data.total_income),
        float(data.total_expense),
        float(data.total_investment),
        float(data.balance),
    ])
    for cell in ws[3]:
        _style_header(cell)
    for cell in ws[4]:
        cell.number_format = 'R$ #,##0.00'
        cell.alignment = Alignment(horizontal="right")
        cell.border = _thin_border
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18

    if data.categories:
        ws.append([])
        ws.append(["Category", "Amount (R$)", "% of expenses"])
        for cell in ws[ws.max_row]:
            _style_header(cell)
        for i, c in enumerate(data.categories):
            ws.append([c.name, float(c.amount), c.percentage / 100])
            row = ws[ws.max_row]
            for cell in row:
                _style_cell(cell, alt=(i % 2 == 1))
            row[1].number_format = 'R$ #,##0.00'
            row[2].number_format = "0.0%"

    # ── Sheet 2: Transactions ─────────────────────────────────────────────
    ws2 = wb.create_sheet("Transactions")
    headers = ["Date", "Description", "Category", "Type", "Amount (R$)"]
    ws2.append(headers)
    for cell in ws2[1]:
        _style_header(cell)
    for i, t in enumerate(data.transactions):
        ws2.append([
            t.date,
            t.description or "",
            t.category,
            t.type.capitalize(),
            float(t.amount),
        ])
        row = ws2[ws2.max_row]
        for cell in row:
            _style_cell(cell, alt=(i % 2 == 1))
        row[4].number_format = 'R$ #,##0.00'

    ws2.column_dimensions["A"].width = 12
    ws2.column_dimensions["B"].width = 35
    ws2.column_dimensions["C"].width = 20
    ws2.column_dimensions["D"].width = 14
    ws2.column_dimensions["E"].width = 16

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Portfolio report data contract ───────────────────────────────────────────

@dataclass
class PositionRow:
    ticker: str
    name: str
    asset_class: str
    quantity: Decimal
    avg_price: Decimal
    total_invested: Decimal
    total_dividends: Decimal


@dataclass
class DividendRow:
    date: str
    ticker: str
    amount: Decimal
    note: str


@dataclass
class AllocationRow:
    asset_class: str
    total_invested: Decimal
    percentage: float


@dataclass
class PortfolioReportData:
    portfolio_name: str
    generated_at: str
    total_invested: Decimal
    total_dividends: Decimal
    positions: list[PositionRow]
    dividends: list[DividendRow]
    allocation: list[AllocationRow]


# ── Portfolio PDF ─────────────────────────────────────────────────────────────

def generate_portfolio_pdf(data: PortfolioReportData) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=2 * cm, rightMargin=2 * cm,
        topMargin=2 * cm, bottomMargin=2 * cm,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("Title", parent=styles["Normal"],
        fontSize=18, textColor=_SLATE_900, spaceAfter=4, fontName="Helvetica-Bold")
    subtitle_style = ParagraphStyle("Subtitle", parent=styles["Normal"],
        fontSize=10, textColor=_SLATE_400, spaceAfter=12)
    section_style = ParagraphStyle("Section", parent=styles["Normal"],
        fontSize=11, textColor=_SLATE_900, spaceBefore=16, spaceAfter=8, fontName="Helvetica-Bold")

    story = []
    story.append(Paragraph("StratEncomy", title_style))
    story.append(Paragraph(f"Portfolio Report — {data.portfolio_name} — {data.generated_at}", subtitle_style))
    story.append(HRFlowable(width="100%", thickness=1, color=_SLATE_700, spaceAfter=16))

    summary_data = [
        ["Total Invested", "Total Dividends", "Positions"],
        [_fmt_safe(data.total_invested), _fmt_safe(data.total_dividends), str(len(data.positions))],
    ]
    st = Table(summary_data, colWidths=["33%"] * 3)
    st.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), _SLATE_900),
        ("TEXTCOLOR",  (0, 0), (-1, 0), _WHITE),
        ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",   (0, 0), (-1, -1), 9),
        ("ALIGN",      (0, 0), (-1, -1), "CENTER"),
        ("FONTNAME",   (0, 1), (-1, 1), "Helvetica-Bold"),
        ("FONTSIZE",   (0, 1), (-1, 1), 11),
        ("TEXTCOLOR",  (0, 1), (0, 1), _GREEN),
        ("TEXTCOLOR",  (1, 1), (1, 1), _BLUE),
        ("ROWBACKGROUNDS", (0, 1), (-1, 1), [colors.HexColor("#F8FAFC")]),
        ("BOX",  (0, 0), (-1, -1), 0.5, _SLATE_700),
        ("GRID", (0, 0), (-1, -1), 0.25, _SLATE_400),
        ("TOPPADDING",    (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
    ]))
    story.append(st)

    if data.allocation:
        story.append(Paragraph("Allocation by Class", section_style))
        alloc_data = [["Class", "Invested", "%"]] + [
            [r.asset_class.upper(), _fmt_safe(r.total_invested), f"{r.percentage:.1f}%"]
            for r in data.allocation
        ]
        at = Table(alloc_data, colWidths=["40%", "40%", "20%"])
        at.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), _SLATE_900),
            ("TEXTCOLOR",  (0, 0), (-1, 0), _WHITE),
            ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE",   (0, 0), (-1, -1), 9),
            ("ALIGN",      (1, 0), (-1, -1), "RIGHT"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [_WHITE, colors.HexColor("#F1F5F9")]),
            ("BOX",  (0, 0), (-1, -1), 0.5, _SLATE_700),
            ("LINEBELOW", (0, 0), (-1, 0), 0.5, _SLATE_700),
            ("TOPPADDING",    (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))
        story.append(at)

    if data.positions:
        story.append(Paragraph("Positions", section_style))
        pos_data = [["Ticker", "Name", "Class", "Qty", "Avg Price", "Invested", "Dividends"]] + [
            [p.ticker, p.name or "—", p.asset_class.upper(),
             f"{float(p.quantity):,.4f}", _fmt_safe(p.avg_price),
             _fmt_safe(p.total_invested), _fmt_safe(p.total_dividends)]
            for p in data.positions
        ]
        pt = Table(pos_data, colWidths=["10%", "20%", "10%", "12%", "14%", "17%", "17%"])
        pt.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), _SLATE_900),
            ("TEXTCOLOR",  (0, 0), (-1, 0), _WHITE),
            ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE",   (0, 0), (-1, -1), 8),
            ("ALIGN",      (3, 0), (-1, -1), "RIGHT"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [_WHITE, colors.HexColor("#F1F5F9")]),
            ("BOX",  (0, 0), (-1, -1), 0.5, _SLATE_700),
            ("LINEBELOW", (0, 0), (-1, 0), 0.5, _SLATE_700),
            ("TOPPADDING",    (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        story.append(pt)

    if data.dividends:
        story.append(Paragraph("Dividends", section_style))
        div_data = [["Date", "Ticker", "Amount", "Note"]] + [
            [d.date, d.ticker, _fmt_safe(d.amount), d.note or "—"]
            for d in data.dividends
        ]
        dt = Table(div_data, colWidths=["15%", "15%", "20%", "50%"])
        dt.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), _SLATE_900),
            ("TEXTCOLOR",  (0, 0), (-1, 0), _WHITE),
            ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE",   (0, 0), (-1, -1), 8),
            ("ALIGN",      (2, 0), (2, -1), "RIGHT"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [_WHITE, colors.HexColor("#F1F5F9")]),
            ("BOX",  (0, 0), (-1, -1), 0.5, _SLATE_700),
            ("LINEBELOW", (0, 0), (-1, 0), 0.5, _SLATE_700),
            ("TOPPADDING",    (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        story.append(dt)

    doc.build(story)
    return buf.getvalue()


# ── Portfolio XLSX ────────────────────────────────────────────────────────────

def generate_portfolio_xlsx(data: PortfolioReportData) -> bytes:
    wb = Workbook()
    _hf  = Font(bold=True, color="F8FAFC", size=10)
    _hfill = PatternFill("solid", fgColor="0F172A")
    _alt = PatternFill("solid", fgColor="F1F5F9")
    _border = Border(
        left=Side(style="thin", color="334155"), right=Side(style="thin", color="334155"),
        top=Side(style="thin", color="334155"), bottom=Side(style="thin", color="334155"),
    )

    def _h(cell):
        cell.font = _hf; cell.fill = _hfill
        cell.alignment = Alignment(horizontal="center", vertical="center"); cell.border = _border

    def _c(cell, alt=False):
        if alt: cell.fill = _alt
        cell.alignment = Alignment(vertical="center"); cell.border = _border

    ws = wb.active
    ws.title = "Positions"
    ws.append([f"StratEncomy — {data.portfolio_name} — {data.generated_at}"])
    ws["A1"].font = Font(bold=True, size=13)
    ws.append([])
    ws.append(["Ticker", "Name", "Class", "Quantity", "Avg Price", "Invested", "Dividends"])
    for cell in ws[3]: _h(cell)
    for i, p in enumerate(data.positions):
        ws.append([p.ticker, p.name or "", p.asset_class.upper(),
                   float(p.quantity), float(p.avg_price), float(p.total_invested), float(p.total_dividends)])
        row = ws[ws.max_row]
        for cell in row: _c(cell, alt=(i % 2 == 1))
        row[4].number_format = 'R$ #,##0.00000'
        row[5].number_format = 'R$ #,##0.00'
        row[6].number_format = 'R$ #,##0.00'
    for col, w in zip("ABCDEFG", [10, 25, 10, 14, 16, 16, 16]):
        ws.column_dimensions[col].width = w

    ws2 = wb.create_sheet("Dividends")
    ws2.append(["Date", "Ticker", "Amount", "Note"])
    for cell in ws2[1]: _h(cell)
    for i, d in enumerate(data.dividends):
        ws2.append([d.date, d.ticker, float(d.amount), d.note or ""])
        row = ws2[ws2.max_row]
        for cell in row: _c(cell, alt=(i % 2 == 1))
        row[2].number_format = 'R$ #,##0.00'
    for col, w in zip("ABCD", [12, 10, 16, 35]):
        ws2.column_dimensions[col].width = w

    ws3 = wb.create_sheet("Allocation")
    ws3.append(["Class", "Invested", "%"])
    for cell in ws3[1]: _h(cell)
    for i, a in enumerate(data.allocation):
        ws3.append([a.asset_class.upper(), float(a.total_invested), a.percentage / 100])
        row = ws3[ws3.max_row]
        for cell in row: _c(cell, alt=(i % 2 == 1))
        row[1].number_format = 'R$ #,##0.00'
        row[2].number_format = "0.0%"
    for col, w in zip("ABC", [14, 18, 12]):
        ws3.column_dimensions[col].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
